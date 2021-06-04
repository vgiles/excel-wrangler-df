from openpyxl import load_workbook
from openpyxl import Workbook
import os
from shutil import copyfile
from openpyxl.styles import PatternFill
import math

path = '.'

subdirs = []

comment_list = []
# Find all directories within a single folder for recursive processing
directory_contents = os.listdir(path)

for item in directory_contents:
    if os.path.isdir(item):
        subdirs.append(item)

print(subdirs)

for x in subdirs:
    paths = [os.path.join(x + '/data', fn) for fn in next(os.walk(x + '/data'))[2]]
    copyfile(paths[0], x + "/" + x + "_.xlsx")
    # print("created a destination file from one of your data files.")

    cell_list=[]
    comment_list=[]
    destination = x + "/" + x + "_moderated.xlsx"
    wbd = load_workbook(destination)
    wsd = wbd.active

    # clear any values that might be in the destination file
    for row in wsd.iter_rows(min_row=13, min_col=3, max_col=17, max_row=17):
        for cell in row:
            if cell.value == 1:
                cell.value = 0
            elif cell.value == None:
                cell.value = 0

    for row in wsd.iter_rows(min_row=19, min_col=3, max_col=17, max_row=20):
        for cell in row:
            if cell.value == 1:
                cell.value = 0
            elif cell.value == None:
                cell.value = 0

    for row in wsd.iter_rows(min_row=22, min_col=3, max_col=17, max_row=23):
        for cell in row:
            if cell.value == 1:
                cell.value = 0
            elif cell.value == None:
                cell.value = 0
    print("File cleared")

    for f in paths:
        source = f
        # print(source)
        wb = load_workbook(source, data_only=True, read_only=True)
        ws = wb.active
        # print(source)
        if ws['A27'].value != None:
            # print(ws['A27'].value)
            comment_list.append(ws['A27'].value)
        # print(comment_list)
        # comment_list = comment_list.append(ws('A27'))
        # Collect cell locations of the value "1" in a single source spreadsheet
        for row in ws.iter_rows(min_row=13, min_col=3, max_col=17, max_row=23):
            for cell in row:
                if cell.value == 1:
                    wsd.cell(cell.row, cell.column).value += 1
                if cell.value == 2:
                    wsd.cell(cell.row, cell.column).value += 1

    wsd['C2'] = len(comment_list)
    if wsd['C2'].value == 3:
        wsd['A27'] = comment_list[0] + "\n" + "\n" + comment_list[1] + "\n" + "\n" + comment_list[2]
    elif wsd['C2'].value == 2:
        wsd['A27'] = comment_list[0] + "\n" + "\n" + comment_list[1]
    elif wsd['C2'].value == 2:
        wsd['A27'] = comment_list[0]


    print("Data merged")

    # print(comment_list)
    wbd.save(destination)
    print(destination + " " + "has been saved")

    # averager

    row_avg = []
    for row in wsd.iter_rows(min_row=13, min_col=3, max_col=17, max_row=17):
        row_accum = []
        for cell in row:
            if cell.value != 0:
                row_accum.append(cell.col_idx - 2)

        z = int(math.ceil(sum(row_accum)/len(row_accum)))+2

        for cell in row:
            if cell.col_idx == z:
                cell.value += 1
                cell.fill = PatternFill(start_color="FFFFAD", end_color="FFFFAD", fill_type = "solid")

    for row in wsd.iter_rows(min_row=19, min_col=3, max_col=17, max_row=20):
        row_accum = []
        for cell in row:
            if cell.value != 0:
                row_accum.append(cell.col_idx - 2)

        z = int(math.ceil(sum(row_accum)/len(row_accum)))+2
        for cell in row:
            if cell.col_idx == z:
                cell.value += 1
                cell.fill = PatternFill(start_color="FFFFAD", end_color="FFFFAD", fill_type = "solid")

    for row in wsd.iter_rows(min_row=22, min_col=3, max_col=17, max_row=23):
        row_accum = []
        for cell in row:
            if cell.value != 0:
                row_accum.append(cell.col_idx - 2)

        z = int(math.ceil(sum(row_accum)/len(row_accum)))+2
        for cell in row:
            if cell.col_idx == z:
                cell.value += 1
                cell.fill = PatternFill(start_color="FFFFAD", end_color="FFFFAD", fill_type = "solid")

    for row in wsd.iter_rows(min_row=13, min_col=3, max_col=17, max_row=23):
        for cell in row:
            if cell.value == 0:
                cell.value = None

    # set marker to + 1 original for moderation calculation
    wsd['C2'] = wsd['C2'].value + 1

    with open(x + "/comments.txt", 'w') as outfile:
        outfile.write(wsd['A27'].value)
    wbd.save(destination)
