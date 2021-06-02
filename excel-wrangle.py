from openpyxl import load_workbook
from openpyxl import Workbook
import os
from shutil import copyfile



paths = [os.path.join('data', fn) for fn in next(os.walk('data'))[2]]
# print(paths)

copyfile(paths[0], "destination.xlsx")

print("created a destination file from one of your data files.")

cell_list=[]
comment_list=[]
destination = "destination.xlsx"
wbd = load_workbook(destination)
wsd = wbd.active

# clear any values that might be in the destination file
for row in wsd.iter_rows(min_row=13, min_col=3, max_col=17, max_row=17):
    for cell in row:
        if cell.value == 1:
            cell.value = 0
        elif cell.value == None:
            cell.value = 0

for row in wsd.iter_rows(min_row=19, min_col=3, max_col=17, max_row=20):
    for cell in row:
        if cell.value == 1:
            cell.value = 0
        elif cell.value == None:
            cell.value = 0

for row in wsd.iter_rows(min_row=22, min_col=3, max_col=17, max_row=23):
    for cell in row:
        if cell.value == 1:
            cell.value = 0
        elif cell.value == None:
            cell.value = 0
print("File cleared")

for f in paths:
    source = f
    # print(source)
    wb = load_workbook(source, data_only=True, read_only=True)
    ws = wb.active
    # print(source)
    comment_list.append(ws['A27'].value)
    # print(comment_list)
    # comment_list = comment_list.append(ws('A27'))
    # Collect cell locations of the value "1" in a single source spreadsheet
    for row in ws.iter_rows(min_row=13, min_col=3, max_col=17, max_row=23):
        for cell in row:
            if cell.value == 1:
                wsd.cell(cell.row, cell.column).value += 1

for row in wsd.iter_rows(min_row=13, min_col=3, max_col=17, max_row=23):
    for cell in row:
        if cell.value == 0:
            cell.value = None

wsd['C2'] = 3
wsd['A27'] = comment_list[0] + "\n" + "\n" + comment_list[1] + "\n" + "\n" + comment_list[2]
print("Data merged")

# print(comment_list)
wbd.save(destination)

print(destination + " " + "has been saved")
